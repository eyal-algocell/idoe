"""Table generation and Excel export for iDoE plans."""

import pandas as pd
import numpy as np
from typing import Dict, List
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment


class TableGenerator:
    """Generates tables and Excel exports for iDoE plans."""

    def __init__(
        self,
        assignments: Dict[int, List[int]],
        combinations: np.ndarray,
        param_names: List[str],
        param_units: List[str],
        total_hours: float,
        num_stages: int,
        constraints_info: Dict = None
    ):
        """
        Initialize table generator.

        Args:
            assignments: Dict mapping run_idx -> [combo_idx per stage]
            combinations: Array of shape (n_combos, n_params)
            param_names: List of parameter names
            param_units: List of parameter units
            total_hours: Total run duration
            num_stages: Number of stages per run
            constraints_info: Dictionary with constraint settings
        """
        self.assignments = assignments
        self.combinations = combinations
        self.param_names = param_names
        self.param_units = param_units
        self.total_hours = total_hours
        self.num_stages = num_stages
        self.per_stage_hours = total_hours / num_stages
        self.constraints_info = constraints_info or {}

    def _format_time_window(self, stage: int) -> str:
        """Format time window string for a stage."""
        t_start = stage * self.per_stage_hours
        t_end = (stage + 1) * self.per_stage_hours
        return f"{t_start:.1f}–{t_end:.1f}"

    def _get_param_columns(self) -> List[str]:
        """Get parameter column names with units."""
        columns = []
        for name, unit in zip(self.param_names, self.param_units):
            if unit:
                columns.append(f"{name} ({unit})")
            else:
                columns.append(name)
        return columns

    def generate_run_table(self, run_idx: int) -> pd.DataFrame:
        """
        Generate table for a single run.

        Args:
            run_idx: Run index (0-based)

        Returns:
            DataFrame with run schedule
        """
        if run_idx not in self.assignments:
            return pd.DataFrame()

        combo_list = self.assignments[run_idx]
        rows = []

        for stage_idx, combo_idx in enumerate(combo_list):
            row = {
                "Run": f"Run {run_idx + 1}",
                "Stage": f"Stage {stage_idx + 1}",
                "Time (h)": self._format_time_window(stage_idx),
                "Combo #": combo_idx + 1  # 1-based
            }

            # Add parameter values
            param_columns = self._get_param_columns()
            for p_idx, col_name in enumerate(param_columns):
                value = self.combinations[combo_idx, p_idx]
                row[col_name] = round(value, 4)

            rows.append(row)

        return pd.DataFrame(rows)

    def generate_combined_table(self) -> pd.DataFrame:
        """
        Generate combined table with all runs.

        Returns:
            DataFrame with all run schedules
        """
        tables = []
        for run_idx in sorted(self.assignments.keys()):
            tables.append(self.generate_run_table(run_idx))

        if not tables:
            return pd.DataFrame()

        return pd.concat(tables, ignore_index=True)

    def generate_constraints_table(self) -> pd.DataFrame:
        """
        Generate table summarizing constraint settings.

        Returns:
            DataFrame with constraint information
        """
        rows = []

        # Parameter definitions
        rows.append({"Setting": "=== Parameters ===", "Value": ""})
        for name, unit in zip(self.param_names, self.param_units):
            param_label = f"{name} ({unit})" if unit else name
            rows.append({"Setting": f"Parameter: {name}", "Value": param_label})

        rows.append({"Setting": "", "Value": ""})

        # Experiment structure
        rows.append({"Setting": "=== Experiment Structure ===", "Value": ""})
        rows.append({"Setting": "Total Runs Used", "Value": len(self.assignments)})
        rows.append({"Setting": "Stages per Run", "Value": self.num_stages})
        rows.append({"Setting": "Run Duration (h)", "Value": self.total_hours})
        rows.append({"Setting": "Stage Duration (h)", "Value": f"{self.per_stage_hours:.1f}"})

        rows.append({"Setting": "", "Value": ""})

        # Constraints
        rows.append({"Setting": "=== Constraints ===", "Value": ""})

        for key, value in self.constraints_info.items():
            rows.append({"Setting": key, "Value": str(value)})

        return pd.DataFrame(rows)

    def create_excel_workbook(self) -> BytesIO:
        """
        Create Excel workbook with multiple sheets.

        Returns:
            BytesIO object containing Excel file
        """
        output = BytesIO()

        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            # Summary sheet - combined table
            combined_df = self.generate_combined_table()
            combined_df.to_excel(writer, sheet_name='Summary', index=False)

            # Per-run sheets
            for run_idx in sorted(self.assignments.keys()):
                run_df = self.generate_run_table(run_idx)
                sheet_name = f'Run_{run_idx + 1}'
                run_df.to_excel(writer, sheet_name=sheet_name, index=False)

            # Constraints sheet
            constraints_df = self.generate_constraints_table()
            constraints_df.to_excel(writer, sheet_name='Constraints', index=False)

        # Style the workbook
        output.seek(0)
        wb = openpyxl.load_workbook(output)

        # Style each sheet
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]

            # Header style
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")

            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")

            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width

        # Save styled workbook
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        return output

    def get_summary_stats(self) -> Dict[str, any]:
        """
        Get summary statistics about the plan.

        Returns:
            Dictionary with summary information
        """
        n_runs = len(self.assignments)
        total_stages_used = sum(len(stages) for stages in self.assignments.values())

        # Count combo usage
        combo_usage = {}
        for combo_list in self.assignments.values():
            for combo_idx in combo_list:
                combo_usage[combo_idx] = combo_usage.get(combo_idx, 0) + 1

        n_combos_used = len(combo_usage)
        max_combo_usage = max(combo_usage.values()) if combo_usage else 0
        avg_combo_usage = sum(combo_usage.values()) / len(combo_usage) if combo_usage else 0

        # Count repeats
        n_repeated_combos = sum(1 for count in combo_usage.values() if count > 1)

        return {
            "runs_used": n_runs,
            "total_stages": total_stages_used,
            "combos_covered": n_combos_used,
            "combos_repeated": n_repeated_combos,
            "max_combo_usage": max_combo_usage,
            "avg_combo_usage": round(avg_combo_usage, 2),
            "total_combos": len(self.combinations),
            "coverage_percent": round(100 * n_combos_used / len(self.combinations), 1)
        }
